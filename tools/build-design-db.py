#!/usr/bin/env python3
"""design/design.json -> SQLite + spreadsheet + a browsable HTML page.

Every table is flattened to columns so it can actually be sorted and filtered.
Nested values (cost arrays, build-bias maps) are kept as JSON text in their own
column rather than dropped: they are rarely what you sort by, but losing them
would make the database a summary instead of a source of truth."""
import json, os, sqlite3, html

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
D = json.load(open(os.path.join(ROOT, 'design/design.json')))
OUT = os.path.join(ROOT, 'design')

def rows_of(name, blob):
    """Normalise both shapes the game uses — arrays and keyed objects — into
    a list of dicts with a stable leading id column."""
    data = blob['data']
    rows = []
    if isinstance(data, list):
        for i, v in enumerate(data):
            r = {'id': i}
            r.update(v if isinstance(v, dict) else {'value': v})
            rows.append(r)
    elif isinstance(data, dict):
        for k, v in data.items():
            r = {'id': k}
            if isinstance(v, dict): r.update(v)
            else: r['value'] = v
            rows.append(r)
    return rows

def flat(v):
    if v is None or isinstance(v, (int, float, str, bool)): return v
    return json.dumps(v, separators=(',', ':'))

tables = {}
for name, blob in D['tables'].items():
    rows = rows_of(name, blob)
    if not rows: continue
    cols, seen = [], set()
    for r in rows:
        for k in r:
            if k not in seen: seen.add(k); cols.append(k)
    tables[name] = (cols, [{c: flat(r.get(c)) for c in cols} for r in rows], blob.get('note', ''))

# ---- SQLite ----------------------------------------------------------------
dbp = os.path.join(OUT, 'massfront-design.db')
if os.path.exists(dbp): os.remove(dbp)
con = sqlite3.connect(dbp); cur = con.cursor()
cur.execute('CREATE TABLE _meta (key TEXT, value TEXT)')
cur.executemany('INSERT INTO _meta VALUES (?,?)',
                [('app_version', D.get('appVersion', '')),
                 ('generated_from', D.get('generatedFrom', '')),
                 ('tables', str(len(tables)))])
for name, (cols, rows, note) in tables.items():
    q = lambda c: '"' + c.replace('"', '') + '"'
    cur.execute('CREATE TABLE %s (%s)' % (name, ','.join(q(c) + ' TEXT' for c in cols)))
    cur.executemany('INSERT INTO %s VALUES (%s)' % (name, ','.join('?' * len(cols))),
                    [[r[c] for c in cols] for r in rows])
    cur.execute('INSERT INTO _meta VALUES (?,?)', ('note:' + name, note))
con.commit(); con.close()

# ---- spreadsheet -----------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
wb = Workbook(); wb.remove(wb.active)
HEAD = PatternFill('solid', fgColor='16232F')
for name, (cols, rows, note) in tables.items():
    ws = wb.create_sheet(name[:31])
    ws.append([note or name]); ws['A1'].font = Font(italic=True, size=9, color='667788')
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = Font(bold=True, color='CFE6FF'); cell.fill = HEAD
        cell.alignment = Alignment(horizontal='left')
    for r in rows: ws.append([r[c] for c in cols])
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = 'A2:%s%d' % (get_column_letter(len(cols)), len(rows) + 2)
    for i, c in enumerate(cols, 1):
        w = max(len(str(c)), *(len(str(r[c] if r[c] is not None else '')) for r in rows)) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(46, max(8, w))
wb.save(os.path.join(OUT, 'massfront-design.xlsx'))

# ---- browsable HTML --------------------------------------------------------
parts = ["""<!doctype html><meta charset=utf-8><title>MASSFRONT design database</title>
<meta name=viewport content="width=device-width,initial-scale=1">
<style>
:root{color-scheme:dark}
body{margin:0;background:#080d15;color:#c8dcec;font:14px/1.55 ui-sans-serif,system-ui,sans-serif}
header{padding:22px 20px;background:linear-gradient(180deg,#0f1c2b,#080d15);border-bottom:1px solid #1d3247}
h1{margin:0;font-size:22px;letter-spacing:.18em;color:#8be8ff}
.sub{color:#7f9db4;font-size:12px;margin-top:5px}
nav{display:flex;flex-wrap:wrap;gap:6px;padding:14px 20px;position:sticky;top:0;background:#0a1119ee;
 backdrop-filter:blur(8px);border-bottom:1px solid #17293a;z-index:5}
nav a{padding:5px 10px;border-radius:7px;background:#132234;color:#9fc6e0;text-decoration:none;font-size:11.5px}
nav a:hover{background:#1c3550}
section{padding:26px 20px}
h2{font-size:15px;letter-spacing:.14em;color:#ffce6a;margin:0 0 3px}
.note{color:#7f9db4;font-size:11.5px;margin-bottom:10px}
.wrap{overflow-x:auto;border:1px solid #17293a;border-radius:10px}
table{border-collapse:collapse;width:100%;font-size:12px}
th{position:sticky;top:0;background:#132234;color:#cfe6ff;text-align:left;padding:8px 10px;
 font-weight:700;white-space:nowrap;border-bottom:1px solid #21384d}
td{padding:6px 10px;border-bottom:1px solid #101c28;vertical-align:top;white-space:nowrap}
tr:nth-child(even) td{background:#0b141d}
td.j{color:#8fb3cc;font-family:ui-monospace,monospace;font-size:11px;white-space:pre-wrap;max-width:340px}
</style>"""]
parts.append('<header><h1>MASSFRONT DESIGN DATABASE</h1><div class=sub>v%s &middot; %d tables &middot; %s</div></header>'
             % (html.escape(D.get('appVersion','')), len(tables), html.escape(D.get('generatedFrom',''))))
parts.append('<nav>' + ''.join('<a href="#%s">%s</a>' % (n, n) for n in tables) + '</nav>')
for name, (cols, rows, note) in tables.items():
    parts.append('<section id="%s"><h2>%s <span style="color:#5d7a90">(%d)</span></h2>' % (name, name.upper(), len(rows)))
    if note: parts.append('<div class=note>%s</div>' % html.escape(note))
    parts.append('<div class=wrap><table><thead><tr>' + ''.join('<th>%s</th>' % html.escape(str(c)) for c in cols) + '</tr></thead><tbody>')
    for r in rows:
        tds = []
        for c in cols:
            v = r[c]; s = '' if v is None else str(v)
            cls = ' class=j' if (s.startswith('{') or s.startswith('[')) else ''
            tds.append('<td%s>%s</td>' % (cls, html.escape(s)))
        parts.append('<tr>' + ''.join(tds) + '</tr>')
    parts.append('</tbody></table></div></section>')
open(os.path.join(OUT, 'design-db.html'), 'w').write('\n'.join(parts))

print('tables : ' + ', '.join('%s(%d)' % (n, len(t[1])) for n, t in tables.items()))
print('wrote  : design/massfront-design.db, massfront-design.xlsx, design-db.html')
